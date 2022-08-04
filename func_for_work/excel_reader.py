from settings import desktop_path
from common_funcs import define_links, define_main_page

file_name = desktop_path + 'Нормирование.xlsx'

import openpyxl

def open_excel(excel_file_name, sheet_number = 0):
    """возвращает генератор инф-и из excel """
    work_book = openpyxl.load_workbook(excel_file_name, read_only = True, data_only = True)
    active_sheet = work_book.worksheets[sheet_number]
    rows_generator = active_sheet.iter_rows(min_row = 2)
    return rows_generator

# читаем excel - рабочую таблицу
def read_work_table(excel_file_name):
    ''' Создает:
        словарь компаний,
        список ссылок
        '''
    rows_generator = open_excel(excel_file_name)
    # номера столбцов (-1)
    comp_inn_clm_num = 16
    comp_name_clm_num = 17

    links_clm_num = 18

    dict_information = {}
    list_links = []
    link_num = 1

    for string_xlsx_row in rows_generator:

        comp_inn = string_xlsx_row[comp_inn_clm_num].value
        comp_name = string_xlsx_row[comp_name_clm_num].value
        if comp_name:
            comp_name = comp_name.upper()
        # link_num = string_xlsx_row[link_num_clm_num].value

        links_list = define_links(string_xlsx_row[links_clm_num].value)
        kkn_part = string_xlsx_row[19].value

        main_page_set = set()

        # сощдание словаря компаний
        if links_list:
            for link in links_list:
                main_page = define_main_page(link)
                main_page_set.add(main_page)
        else:
            # print(f"№ {link_num} - пропускаем")
            continue

        if "zakupki.gov.ru" in main_page_set:
            # Эти сайты не обрабатываем
            continue

        if comp_inn:
            if comp_inn in dict_information:
                dict_information[comp_inn]["main_pages"] = dict_information[comp_inn]["main_pages"] | main_page_set
                if kkn_part:
                    dict_information[comp_inn]["kkn_part"].add(kkn_part)

                if dict_information[comp_inn]["comp_name"] != comp_name:
                    print(f"Не совпадает {comp_inn}: {comp_name}")
            else:
                dict_information[comp_inn] = {"comp_name": comp_name, "main_pages": main_page_set, 'kkn_part': {kkn_part}}

        # Создание списка ссылок
        if len(links_list) > 1:

            cntr = 1
            for link in links_list:
                # list_links.append([comp_inn, main_page, f"{link_num}_0{cntr}", links_list[0],'' ,'' , kkn_part])
                list_links.append([comp_inn, main_page, int(link_num), links_list[0],'' ,'' , kkn_part])
                link_num += 1
                cntr += 1
        else:
            list_links.append([comp_inn, main_page, int(link_num), links_list[0],'' ,'' , kkn_part]) #str(link_num)

        link_num += 1

    return dict_information, list_links

#
def read_links_table(file_path):
    links_table_gen = open_excel(file_path)

    links_set = set()
    num_set = set()
    out_list = []
    for row in links_table_gen:
        current_values = []
        for r_value in row:
            current_values.append(r_value.value)
        out_list.append(current_values)

        links_set.add(row[3].value)
        num_set.add(row[2].value)

        # print(row[2].value, row[3].value)
    if num_set:
        max_number = max(num_set)
    else:
        max_number = 1
    return max_number, links_set, out_list

def update_companies_file(file_path, new_comp_data):
    # Открываем файл
    wb = openpyxl.load_workbook(file_path)
    cush = wb['companies_info']
    pars_sheet = wb['site_settings']
    appended_rows = []
    new_main_page_dict = dict()

    curren_inn_set = {inn_val.value for inn_val in cush['A']}
    for inn_val, comp_info in new_comp_data.items():
        # Поиск строки по инн
        new_row = []
        # создаем заодно словарь главных страниц для 2 листа
        for main_page in comp_info['main_pages']:
            if main_page not in new_main_page_dict:
                new_main_page_dict[main_page] = inn_val

        if inn_val not in curren_inn_set:
            new_row = [inn_val, comp_info['comp_name'], '', '', ";".join(comp_info['kkn_part'])]
            # print(new_row)
            appended_rows.append(new_row)
        else:
            for inn_clmn in cush['A']:
                if inn_val == inn_clmn.value:
                    # проверка найденной:
                    # на совпадение названия компании
                    if comp_info['comp_name'] != cush[f'B{cush["A"].index(inn_clmn) + 1}'].value:
                        print('Не совпадение по названию:')
                        print([comp_info['comp_name']], "\n", [cush[f'B{cush["A"].index(inn_clmn) + 1}'].value])
                        # предупреждение
                    # на наличие частей
                    parts_cel = cush[f'E{cush["A"].index(inn_clmn) + 1}']
                    parts_set = set(parts_cel.value.split(";"))
                    parts_set = parts_set | comp_info['kkn_part']
                    parts_cel.value = ";".join(parts_set)
                        # добавление отсутствующей части

    # дозапись данных, которые не найдены по инн
    for row in appended_rows:
        cush.append(row)

    # дозаписываем новый сайты на второй лист
    appended_main_page_rows = []
    currenе_page_set = {page.value for page in pars_sheet['B']}
    for new_main_page in new_main_page_dict:
        if new_main_page not in currenе_page_set:
            appended_main_page_rows.append([new_main_page_dict[new_main_page], new_main_page])

    for row in appended_main_page_rows:
        pars_sheet.append(row)

    wb.save(file_path)
