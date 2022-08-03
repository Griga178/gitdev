import openpyxl
import os


def create_new_link_table(file_path):
    wb = openpyxl.Workbook()

    current_sheet = wb.active

    current_sheet.title = "work_links"

    current_sheet.cell(row = 1, column = 1).value = "Company INN"
    current_sheet.column_dimensions['A'].width = 15
    current_sheet.cell(row = 1, column = 2).value = "Main page"
    current_sheet.column_dimensions['B'].width = 13
    current_sheet.cell(row = 1, column = 3).value = "Link/screen number"
    current_sheet.column_dimensions['C'].width = 19
    current_sheet.cell(row = 1, column = 4).value = "Link"
    current_sheet.column_dimensions['D'].width = 19
    current_sheet.cell(row = 1, column = 5).value = "Price"
    current_sheet.column_dimensions['E'].width = 13
    current_sheet.cell(row = 1, column = 6).value = "Name"
    current_sheet.column_dimensions['F'].width = 13
    current_sheet.cell(row = 1, column = 7).value = "Parsed date"
    current_sheet.column_dimensions['G'].width = 13
    current_sheet.cell(row = 1, column = 7).value = "Status"
    current_sheet.column_dimensions['H'].width = 13
    current_sheet.cell(row = 1, column = 7).value = "Part"
    current_sheet.column_dimensions['I'].width = 13

    wb.save(file_path)

def create_companies_file(file_path):

    wb = openpyxl.Workbook()

    # Создание первого листа
    current_sheet = wb.active

    current_sheet.title = "companies_info"
    current_sheet.cell(row = 1, column = 1).value = "Company INN"
    current_sheet.column_dimensions['A'].width = 15
    current_sheet.cell(row = 1, column = 2).value = "Name"
    current_sheet.column_dimensions['B'].width = 13
    current_sheet.cell(row = 1, column = 3).value = "Addres"
    current_sheet.column_dimensions['C'].width = 19
    current_sheet.cell(row = 1, column = 4).value = "Phone"
    current_sheet.column_dimensions['D'].width = 13

    # Создание второго листа
    current_sheet_2 = wb.create_sheet("site_settings")
    current_sheet_2.cell(row = 1, column = 1).value = "Company INN"
    current_sheet_2.column_dimensions['A'].width = 15
    current_sheet_2.cell(row = 1, column = 2).value = "Main page"
    current_sheet_2.column_dimensions['B'].width = 19
    current_sheet_2.cell(row = 1, column = 3).value = "Tag"
    current_sheet_2.column_dimensions['C'].width = 13
    current_sheet_2.cell(row = 1, column = 4).value = "Attribute"
    current_sheet_2.column_dimensions['D'].width = 13
    current_sheet_2.cell(row = 1, column = 5).value = "Attribute value"
    current_sheet_2.column_dimensions['E'].width = 15
    current_sheet_2.cell(row = 1, column = 6).value = "Tag type"
    current_sheet_2.column_dimensions['F'].width = 13


    wb.save(file_path)

def check_companies_info_file(file_path):
    file_exists = os.path.isfile(file_path)
    if file_exists:
        print(f'Найден файл: "{file_path.split("/")[-1]}"')
    else:
        create_companies_file(file_path)
        print(f'"{file_path.split("/")[-1]}" - создан')

def check_links_table_file(file_path):
    file_exists = os.path.isfile(file_path)
    if file_exists:
        print(f'Найден файл: "{file_path.split("/")[-1]}"')
    else:
        create_new_link_table(file_path)
        print(f'"{file_path.split("/")[-1]}" - создан')

def update_link_table(excel_info):
    pass


def update_companies_file():
    pass
