'''
    читаем ссылки
'''
from openpyxl import Workbook, load_workbook
from o_1 import get_seller_id

path_desktop = 'C:/Users/G.Tishchenko/Desktop/'

xl_name = path_desktop + 'ozon.xlsx'

def check_link(link):

    split_content = link.split('/')
    domain = split_content[2]
    if domain == 'www.ozon.ru':
        split_content_2 = link.split('?')
        return split_content_2[0]
    else:
        return 'error_link'

def main(file_name):
    counter = 0
    wb = load_workbook(file_name)

    active_sheet = wb.active


    for active_row in active_sheet:
        counter += 1
        print(counter)
        column_num = active_row[0].column
        row_num = active_row[0].row

        current_link = active_row[0].value
        if len(active_row) == 1:
            ''' Строка не обрабатывалась '''
            clear_link = check_link(current_link)
            ''' Проверка/Очистка ссылки'''
            if clear_link:
                value_2 = clear_link
                active_sheet.cell(row = row_num, column = 2).value = value_2
            else:
                value_2 = 'Что то не так с ссылкой'
                print(value_2, current_link)
                active_sheet.cell(row = row_num, column = 2).value = None
                wb.save(xl_name)
                continue

            ''' Парсим id продавца '''
            try:
                seller_id, company_name_1 = get_seller_id(clear_link)
                active_sheet.cell(row = row_num, column = 3).value = seller_id
                active_sheet.cell(row = row_num, column = 4).value = company_name_1
            except:
                seller_id = company_name_1 = 'parse_error'
                active_sheet.cell(row = row_num, column = 3).value = seller_id
                active_sheet.cell(row = row_num, column = 4).value = company_name_1
                wb.save(xl_name)
                continue

        elif len(active_row) == 2:
            ''' В строке обработана только ссылка '''
            clear_link = active_row[1].value
            if clear_link:
                value_2 = clear_link
                active_sheet.cell(row = row_num, column = 2).value = value_2
            else:
                value_2 = 'Что то не так с ссылкой'
                print(value_2, current_link)
                active_sheet.cell(row = row_num, column = 2).value = None
                wb.save(xl_name)
                continue

            ''' Парсим id продавца '''
            try:
                seller_id, company_name_1 = get_seller_id(clear_link)
                active_sheet.cell(row = row_num, column = 3).value = seller_id
                active_sheet.cell(row = row_num, column = 4).value = company_name_1
            except:
                seller_id = company_name_1 = 'parse_error'
                active_sheet.cell(row = row_num, column = 3).value = seller_id
                active_sheet.cell(row = row_num, column = 4).value = company_name_1
                wb.save(xl_name)
                continue



        elif len(active_row) == 4:
            ''' Ссылка парсилась '''
            seller_id = active_row[2].value
            company_name_1 = active_row[3].value
            if seller_id == 'parse_error':
                print("Парсим имя 2 и ОГРН")
            else:
                print("Опять ошибка")
                continue

        else:
            continue








        active_sheet.cell(row = row_num, column = 3).value = seller_id

        active_sheet.cell(row = row_num, column = 5).value = company_name_2
        active_sheet.cell(row = row_num, column = 6).value = company_ogrn
        active_sheet.cell(row = row_num, column = 7).value = address


        wb.save(xl_name)

main(xl_name)
