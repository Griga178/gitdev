# read_file.py
from openpyxl import Workbook, load_workbook

'''
    Порядок столбцов в файле:
    1 ссылка (не исп.)
    2 новое название скрина (только номер)
    3 старое название скрина
    4 цена (не исп.)
'''

def read_file(name: str, sheet_name: str = 'Лист1') -> list[dict]:
    """
    Читает Excel-файл и возвращает список словарей с именами файлов.

    Параметры:
        name (str): Путь к файлу Excel.
        sheet_name (str, optional): Название листа для чтения. По умолчанию 'Лист1'.

    Возвращает:
        list[dict]: Список словарей вида {'old_name': значение_столбца_3, 'new_name': значение_столбца_2}.
                    Данные извлекаются начиная со второй строки до конца.

    Предполагается, что в таблице:
        - столбец 2 (индекс 1) содержит новое имя файла (new_name),
        - столбец 3 (индекс 2) содержит старое имя файла (old_name).
    """

    wb = load_workbook(name, read_only = True, data_only = True)
    active_sheet = wb[sheet_name]
    return_list = []
    for row in active_sheet.iter_rows(min_row = 2, values_only = True):

        return_list.append(
            {
            'old_name': row[2],
            'new_name': row[1],
            }
        )

    return return_list
