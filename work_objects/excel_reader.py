from openpyxl import load_workbook
from typing import List

from .reestr_object import Reestr
# from .directory_object import Directory

class Excel():

    def __init__(self, file_path):
        self.file_path = file_path
        self.columns = []

        self.reestr = Reestr(self.columns, self.get_rows())
        self.directory = False # в случае загрузки Справочника ККН

    def get_rows(self, **kwargs) -> List[List]:
        '''
            kwargs:
                sheet_name: str = False
                header_rows: int = 0 - добавляет в return первую строку
                headers_names: List[str, ...] - возвращяет только указанные
                    колонки, если есть
        '''
        wb = load_workbook(self.file_path, read_only = True, data_only = True)

        sheet_name = kwargs['sheet_name'] if 'sheet_name' in kwargs else wb.sheetnames[0]

        min_row = kwargs['header_rows'] if kwargs.get('header_rows', False) else 2

        active_sheet = wb[sheet_name]
        column_indexes = []
        if kwargs.get('headers_names'):
            headers = [head_name.upper() for head_name in kwargs['headers_names']]
            for column_name in active_sheet[1]:
                if column_name.value.upper() in headers:
                    column_index = active_sheet[1].index(column_name)
                    self.columns.append(column_name.value)
                    column_indexes.append(column_index)
        else:
            for column_name in active_sheet[1]:
                self.columns.append(column_name.value)

        return_list = []
        for row in active_sheet.iter_rows(min_row = min_row, values_only = True):
            return_row = []
            if column_indexes:
                for clm_index in column_indexes:
                    return_row.append(row[clm_index])
            else:
                return_row = list(row)
            return_list.append(return_row)

        return return_list
