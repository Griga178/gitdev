from openpyxl import Workbook, load_workbook
from o_1 import get_seller_id
from o_4 import get_name_ogrn
from o_5 import get_info

path_desktop = 'C:/Users/G.Tishchenko/Desktop/'

xl_name = path_desktop + 'ozon.xlsx'

def check_link(link):

    split_content = link.split('/')
    domain = split_content[2]
    if domain == 'www.ozon.ru':
        split_content_2 = link.split('?')
        return split_content_2[0]
    else:
        return 'error_link'

def main(file_name):
    counter = 0
    wb = load_workbook(file_name)
    active_sheet = wb.active

    for active_row in active_sheet.iter_rows(2):
        # if counter >= 3:
        #     continue

        counter += 1
        row_is_changed = False
        print(counter)
        column_num = active_row[0].column
        row_num = active_row[0].row

        current_link = active_row[0].value

        row_dict = {
            'dirty_link': current_link,
            'link': False,
            'seller_id': False,
            'company_name_1': False,
            'company_name_2': False,
            'company_ogrn': False,
            'company_address': False,
        }

        if len(active_row) == 1:
            row_dict['link'] = check_link(current_link)
            row_is_changed = True
            # Обработка ссылки
            # ЕСЛИ ССЫЛКА == ОК
            # Парсинг id и имя 1
            # Парсинг имя 2, огрн и адрес
        elif len(active_row) == 2:
            row_dict['link'] = active_row[1].value
            # ЕСЛИ ССЫЛКА == ОК
            # Парсинг id и имя 1
            # Парсинг имя 2, огрн и адрес
        elif len(active_row) == 4:
            row_dict['link'] = active_row[1].value
            row_dict['seller_id'] = active_row[2].value
            row_dict['company_name_1'] = active_row[3].value
            # Парсинг имя 2, огрн и адрес


        else:
            comp_name = active_row[4].value
            row_dict['link'] = active_row[1].value
            row_dict['seller_id'] = active_row[2].value
            row_dict['company_name_1'] = active_row[3].value
            row_dict['company_name_2'] = active_row[4].value


        if row_dict['link'] != 'error_link':
            # Если есть проблемы с seller_id - пробуем еще
            if row_dict['seller_id'] == False or row_dict['seller_id'] == 'parse_error':
                print('Вышли сюда 000', row_dict['seller_id'])
                try:
                    row_dict['seller_id'], row_dict['company_name_1'] = get_seller_id(row_dict['link'])
                    row_is_changed = True
                except:
                    row_dict['seller_id'] = row_dict['company_name_1'] = 'parse_error'

            else:
                print('Вышли сюда 111')

            if row_dict['seller_id'] != 'parse_error':
                if not row_dict['company_name_2']:
                    try:
                        c_name_2, c_ogrn, c_address = get_name_ogrn(row_dict['seller_id'])
                        row_is_changed = True
                        row_dict['company_name_2'] = c_name_2
                        row_dict['company_ogrn'] = c_ogrn
                        row_dict['company_address'] = c_address
                    except:
                        print("ошибка парсинга имя 2 и ОГРН")
                        c_name_2 = c_ogrn = c_address = None
                        print(c_name_2, c_ogrn, c_address)
                else:
                    print('Вышли сюда 222')
            else:
                print('Вышли сюда 333')


        if row_is_changed:
            active_sheet.cell(row = row_num, column = 2).value = row_dict['link']
            active_sheet.cell(row = row_num, column = 3).value = row_dict['seller_id']
            active_sheet.cell(row = row_num, column = 4).value = row_dict['company_name_1']
            active_sheet.cell(row = row_num, column = 5).value = row_dict['company_name_2']
            active_sheet.cell(row = row_num, column = 6).value = row_dict['company_ogrn']
            active_sheet.cell(row = row_num, column = 7).value = row_dict['company_address']
            wb.save(xl_name)

        # print(active_sheet.cell(row = row_num, column = 8).value)
        # egrul_parse = active_row[8].value
        if active_sheet.cell(row = row_num, column = 8).value == False:
            print('На ЕГРЮЛ')
            ogrn_val = active_sheet.cell(row = row_num, column = 7).value
            if ogrn_val != False:
                print(ogrn_val)
                try:
                    egrul_info = get_info(ogrn_val)
                    print(egrul_info)
                    active_sheet.cell(row = row_num, column = 8).value = True
                    active_sheet.cell(row = row_num, column = 9).value = egrul_info['full_name']
                    active_sheet.cell(row = row_num, column = 10).value = egrul_info['inn']
                    active_sheet.cell(row = row_num, column = 11).value = egrul_info['address']
                    active_sheet.cell(row = row_num, column = 12).value = egrul_info['date']
                    active_sheet.cell(row = row_num, column = 13).value = egrul_info['kpp']
                    active_sheet.cell(row = row_num, column = 14).value = egrul_info['name']
                    active_sheet.cell(row = row_num, column = 15).value = egrul_info['manager']
                    wb.save(xl_name)
                except:
                    print('Неудачный ЕГРЮЛ')
                    print(ogrn_val)

main(xl_name)
