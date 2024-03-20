'''
    Сохраняем ссылки в БД

'''
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import sessionmaker
from sqlalchemy.orm.exc import NoResultFound, MultipleResultsFound

from database import *

DBSession = sessionmaker(bind = engine)
db_session = DBSession()

def check_link(link):

    split_content = link.split('/')
    domain = split_content[2]
    if domain == 'www.ozon.ru':
        split_content_2 = link.split('?')
        return split_content_2[0]# + '?oos_search=false'
    else:
        return False

path_desktop = 'C:/Users/G.Tishchenko/Desktop/'

excel_name = path_desktop + 'ozon_2.xlsx'

def save_all_links():
    wb = load_workbook(excel_name)
    active_sheet = wb.active

    for active_row in active_sheet.iter_rows(2):

        excel_link = active_row[0].value
        clean_link = check_link(excel_link)
        # print(clean_link)

        if clean_link:

            query = db_session.query(Link)
            query = query.filter_by(**{'clear_value': clean_link})

            try:
                link_obj = query.one()
                print('Уже есть')
            except NoResultFound:
                link_obj = Link(**{
                    'excel_value':excel_link,
                    'clear_value':clean_link,
                    }
                )
                db_session.add(link_obj)
                db_session.commit()
                print(f'Добавлен: {link_obj.clear_value}')
