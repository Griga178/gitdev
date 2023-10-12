from openpyxl import Workbook, load_workbook

def list_to_excel(py_list, excel_path):
    wb = Workbook()
    current_sheet = wb.active
    for el in py_list:
        current_sheet.append(el)
    wb.save(excel_path)

def excel_to_list(xl_name, **kwargs):
    wb = load_workbook(xl_name, read_only = True, data_only = True)
    sheet_name = kwargs['sheet_name'] if 'sheet_name' in kwargs else wb.sheetnames[0]
    min_row = 2 if kwargs.get('headers') == False else 1
    active_sheet = wb[sheet_name]
    column_indices = []
    if kwargs.get('headers_names'):
        for column_name in active_sheet[1]:
            if column_name.value in kwargs['headers_names']:
                column_index = active_sheet[1].index(column_name)
                column_indices.append(column_index)
                kwargs['headers_names'].remove(column_name.value)
        else:
            if kwargs["headers_names"]:
                print(f'НЕ НАШЛОСЬ КОЛОНКА(И) {kwargs["headers_names"]}')
    # headers_indexes =
    return_list = []
    for row in active_sheet.iter_rows(min_row = min_row, values_only = True):
        return_row = []
        if column_indices:
            for clm_index in column_indices:
                return_row.append(row[clm_index])
        else:
            return_row = list(row)
        return_list.append(return_row)

    return return_list

def excel_to_dicts(xl_name, **kwargs):
    '''
        [
            {1:'text', 2:'text2'}, # нет headers_names
            {'name1':} # есть headers_names
        ]
    '''
    wb = load_workbook(
        xl_name,
        read_only = kwargs.get('read_only', True),
        data_only = kwargs.get('data_only', True)
        )

    sheet_name = kwargs.get('sheet_name', wb.sheetnames[0])

    min_row = 2 if kwargs.get('headers') == False else 1

    active_sheet = wb[sheet_name]

    column_indices = []    # {"key":"name" or index}


    if kwargs.get('headers_names'):
        # выбираем столбцы, добавляемые в вывод
        for column_name in active_sheet[1]: # Первая строка с заголовками
            if column_name.value in kwargs['headers_names']:
                # номер столбца
                column_index = active_sheet[1].index(column_name)
                c_key = column_index
                c_val = column_name.value
                column_indices.append({
                    'key': c_key,
                    'value': c_val,
                    })
                kwargs['headers_names'].remove(column_name.value)
        else:
            if kwargs["headers_names"]:
                print(f'НЕ НАШЛОСЬ КОЛОНКА(И) {kwargs["headers_names"]}')
    else:
        # Добавляем все столбцы: ключ = index столбца
        ind = 0
        for column_name in active_sheet[1]:
            column_indices.append({
                'key': ind,
                'value': ind,
                })
            ind += 1

    return_dicts = []
    for row in active_sheet.iter_rows(min_row = min_row, values_only = True):
        return_row = {}
        for clm_index in column_indices:

            row_key = clm_index['value']
            row_val = row[clm_index['key']]
            return_row[row_key] = row_val

        return_dicts.append(return_row)

    return return_dicts
