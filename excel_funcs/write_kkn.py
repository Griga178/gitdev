# write_kkn.py
# запись ккн в отформатированный справочник ККН
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font


def write_worked_list_to_excel(worked_list: list, updated_ktru_excel_path: str) -> None:
    """
    worked_list = [kkn_dict, ...]

    kkn_dict - словарь с ключами и типами значений:
        num: int
        kkn_name: str
        kkn_okpd_2: str
        kkn_det_num: str
        kkn_unit: str
        kkn_chars: list[kkn_char, ...]
        category_name: str
        ktru_number: str
        kkn_number: str
        product_part: str
        upt_date: str
        is_russian: str
        rrrp_number: str

    kkn_char - словарь с ключами и типами значений:
        name: str
        unit: str
        value_is_range: bool
        value_range: tuple[float, float]  # (min, max)
        values: list[str, ...]
        ktru_data: ktru_data_dict

    ktru_data_dict - словарь с ключами и типами значений:
        ktru_char: bool
        all_right: bool
        errors: list[str, ...]
        is_required: bool
        ktru_chars_origin: list[str, ...]
    """
    ws_name = "NewSheet"
    wb = openpyxl.load_workbook(updated_ktru_excel_path)
    ws = wb.create_sheet(ws_name)

    # Рисуем шапку
        # I ряд
    row_1 = ["№", "Наименование ККН", "ОКПД2", "Детализация", "Единица измерения ККН",
    "Показатель (характеристика) товара", "Требования к значениям показателей",
    "", "", "", "Единица измерения характеристики", "Категория", "Код КТРУ",
    "Код ККН", "Товарная часть", "Дата актуализации", "Российский товар", "РРПП"]
    for c_idx, val in enumerate(row_1):
        ws.cell(row=1, column=c_idx+1, value=val)
        # II ряд
    row_2_1 = "Минимальное значение показателя и/или максимальное значение показателя"
    row_2_2 = "Показатели (характеристики), для которых указаны варианты значений"
    row_2_3 = "Показатели, (характеристики) значения которых не могут изменяться"
    ws.cell(row=2, column=7, value=row_2_1)
    ws.cell(row=2, column=9, value=row_2_2)
    ws.cell(row=2, column=10, value=row_2_3)
        # III ряд
    row_3_1 = "≥ (не менее)"
    row_3_2 = "≤ (не более)"
    ws.cell(row=3, column=7, value=row_3_1)
    ws.cell(row=3, column=8, value=row_3_2)
        # IIII ряд
    for i in range(1, 19):
        ws.cell(row=4, column=i, value=i)
        ws.cell(row=4, column=i).alignment = Alignment(horizontal='center')

    # задаем ширину столбцов
    clm_param = {"A": 10, "B": 30, "C": 12, "D": 12, "E": 12, "F": 30,
    "G": 8, "H": 8, "I": 25, "J": 25, "K": 12, "L": 25,
    "M": 25, "N": 25, "O": 25, "P": 10, "Q": 10, "R": 10}
    for clm, width in clm_param.items():
        ws.column_dimensions[clm].width = width

    # Объединяем ячейки
        # 1 строка
    ws.merge_cells(start_row=1, start_column=7,end_row=1, end_column=10)
    ws.cell(row=1, column=7).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    merged_cals = [1, 2,3 , 4, 5, 6, 11, 12, 13, 14, 15, 16, 17, 18] # используется по всей таблицу
    for i in merged_cals:
        ws.merge_cells(start_row=1, start_column=i,end_row=3, end_column=i)
        ws.cell(row=1, column=i).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    merged_cals.remove(6); merged_cals.remove(11)
    merged_cals.remove(15); merged_cals.remove(16)

        # 2 строка
    ws.merge_cells(start_row=2, start_column=7,end_row=2, end_column=8)
    ws.cell(row=2, column=7).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.merge_cells(start_row=2, start_column=9,end_row=3, end_column=9)
    ws.cell(row=2, column=9).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.merge_cells(start_row=2, start_column=10,end_row=3, end_column=10)
    ws.cell(row=2, column=10).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # Установка автофильтра на первую строку (заголовки)
    ws.auto_filter.ref = "A4:R4"

    # заполняем таблицу данными
    main_keys = {'num': 1, 'kkn_name': 2, 'kkn_okpd_2': 3, 'kkn_det_num': 4, 'kkn_unit': 5,
                 'category_name': 12, 'ktru_number': 13, 'kkn_number': 14, 'product_part': 15,
                 'upt_date': 16, 'rrrp_number': 18}

    current_row = 5
    kkn_chars_start_col = 6
    for item in worked_list:
        kkn_chars = item.get("kkn_chars", [])
        chars_len = len(kkn_chars)
        # Записываем основные поля (в 1-й строке записи)
        for key, col in main_keys.items():
            val = item.get(key, "")
            # Если значение не строка, преобразуем в строку
            if not isinstance(val, str):
                val = str(val)
            ws.cell(row=current_row, column=col, value=val)

        is_rus_str = "Да" if item['is_russian'] else ""
        ws.cell(row=current_row, column=17, value=is_rus_str)

        for i in range(chars_len):
            row_i = current_row + i
            if i < len(kkn_chars):
                char = kkn_chars[i]
                ws.cell(row=row_i, column=6, value=char.get("name", ""))
                vr_min = " "
                vr_max = " "
                value_str_9 = " "
                value_str_10 = " "
                if char.get("value_is_range"):
                    vr_min, vr_max = char.get("value_range", (None, None))
                    vr_min = str(vr_min)
                    vr_max = str(vr_max)

                else:
                    values = char.get("values", [])
                    if len(values) > 1:
                        value_str_10 = "; ".join(values)

                    else:
                        value_str_9 = values[0]

                ws.cell(row=row_i, column=7, value=vr_min)
                ws.cell(row=row_i, column=8, value=vr_max)
                ws.cell(row=row_i, column=9, value=value_str_9)
                ws.cell(row=row_i, column=10, value=value_str_10)
                unit_str = char.get("unit", "-")
                ws.cell(row=row_i, column=11, value=unit_str).alignment = Alignment(horizontal='center')

                ws.cell(row=row_i, column=kkn_chars_start_col + 9, value=item.get("product_part", ""))
                date_str = item.get("upt_date", "")
                date_str = date_str[:10] if date_str else ""
                date_str = date_str.replace("-", ".")
                ws.cell(row=row_i, column=kkn_chars_start_col + 10, value=date_str)
            else:           # Если характеристик меньше max - оставляем пустые или "**"
                for c_offset in range(6):
                    ws.cell(row=row_i, column=kkn_chars_start_col + c_offset, value="").font = font

        # Объединяем ячейки с основными полями по высоте, равной количеству строк характеристик
        for col in merged_cals:
            ws.merge_cells(start_row=current_row, start_column=col,
                           end_row=current_row + chars_len-1, end_column=col)
            ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Переходим к следующей записи
        current_row += chars_len

    # устанавливаем границы и параметры шрифта всей таблицы
    font = Font(name='Times New Roman', size=10)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=18):
        for cell in row:
            cell.border = border
            cell.font = font

    wb.save(updated_ktru_excel_path)
