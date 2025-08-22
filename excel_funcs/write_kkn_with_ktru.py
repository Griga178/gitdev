
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill


def write_checked_worked_list_to_excel(worked_list: list, updated_ktru_excel_path: str) -> None:
    """
    worked_list = [kkn_dict, ...]

    kkn_dict - словарь с ключами и типами значений:
        num: int
        kkn_name: str
        kkn_okpd_2: str
        kkn_det_num: str
        kkn_unit: str
        kkn_chars: list[kkn_char, ...]
        category_name: str
        ktru_number: str
        kkn_number: str
        product_part: str
        upt_date: str
        is_russian: str
        rrrp_number: str

    kkn_char - словарь с ключами и типами значений:
        name: str
        unit: str
        value_is_range: bool
        value_range: tuple[float, float]  # (min, max)
        values: list[str, ...]
        ktru_data: ktru_data_dict

    ktru_data_dict - словарь с ключами и типами значений:
        ktru_char: bool
        all_right: bool
        errors: list[str, ...]
        is_required: bool
        ktru_chars_origin: list[str, ...]
    """
    # Желтый цвет
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # Оранжевый цвет
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    # Зеленый цвет
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    # Красный цвет
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    ws_name = "NewSheet"
    wb = openpyxl.load_workbook(updated_ktru_excel_path)
    ws = wb.create_sheet(ws_name)

    # Рисуем шапку
        # I ряд
    row_1 = ["№", "Наименование ККН", "ОКПД2", "Детализация", "Единица измерения ККН",
    "Показатель (характеристика) товара", "Требования к значениям показателей",
    "", "", "", "Единица измерения характеристики", "Категория", "Код КТРУ",
    "Код ККН", "Товарная часть", "Дата актуализации", "Российский товар", "РРПП",
    "Характеристика КТРУ", "Значений нет в КТРУ (вид может отличаться от оригинала)",
     'Все значения КТРУ', 'ед.изм. КТРУ','Статус КТРУ', 'добавленные значения','удаленные значения']
    for c_idx, val in enumerate(row_1):
        ws.cell(row=1, column=c_idx+1, value=val)
        # II ряд
    row_2_1 = "Минимальное значение показателя и/или максимальное значение показателя"
    row_2_2 = "Показатели (характеристики), для которых указаны варианты значений"
    row_2_3 = "Показатели, (характеристики) значения которых не могут изменяться"
    ws.cell(row=2, column=7, value=row_2_1)
    ws.cell(row=2, column=9, value=row_2_2)
    ws.cell(row=2, column=10, value=row_2_3)
        # III ряд
    row_3_1 = "≥ (не менее)"
    row_3_2 = "≤ (не более)"
    ws.cell(row=3, column=7, value=row_3_1)
    ws.cell(row=3, column=8, value=row_3_2)
        # IIII ряд
    for i in range(1, 26):
        ws.cell(row=4, column=i, value=i)
        ws.cell(row=4, column=i).alignment = Alignment(horizontal='center')

    # задаем ширину столбцов
    clm_param = {"A": 8, "B": 15, "C": 12, "D": 8, "E": 8, "F": 30,
    "G": 8, "H": 8, "I": 25, "J": 25, "K": 12, "L": 14,
    "M": 12, "N": 12, "O": 8, "P": 10, "Q": 6, "R": 5,
    "S": 25, "T": 25, "U": 20, "V": 16, "W": 10, "X": 25, "Y": 25}

    for clm, width in clm_param.items():
        ws.column_dimensions[clm].width = width

    # Объединяем ячейки
        # 1 строка
    ws.merge_cells(start_row=1, start_column=7,end_row=1, end_column=10)
    ws.cell(row=1, column=7).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    merged_cals = [1, 2,3 , 4, 5, 12, 13, 14, 17, 18] # используется по всей таблицу
    ktru_clmns = [6, 11, 15, 16, 19, 20, 21, 22, 23, 24, 25]
    for i in merged_cals + ktru_clmns:
        ws.merge_cells(start_row=1, start_column=i,end_row=3, end_column=i)
        ws.cell(row=1, column=i).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # 2 строка
    ws.merge_cells(start_row=2, start_column=7,end_row=2, end_column=8)
    ws.cell(row=2, column=7).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.merge_cells(start_row=2, start_column=9,end_row=3, end_column=9)
    ws.cell(row=2, column=9).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.merge_cells(start_row=2, start_column=10,end_row=3, end_column=10)
    ws.cell(row=2, column=10).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # Установка автофильтра на первую строку (заголовки)
    ws.auto_filter.ref = "A4:X4"

    # заполняем таблицу данными
    main_keys = {'num': 1, 'kkn_name': 2, 'kkn_okpd_2': 3, 'kkn_det_num': 4, 'kkn_unit': 5,
                 'category_name': 12, 'ktru_number': 13, 'kkn_number': 14, 'product_part': 15,
                 'upt_date': 16, 'rrrp_number': 19}

    current_row = 5
    kkn_chars_start_col = 6

    print_stage_on = True
    item_lenght = len(worked_list)
    item_idx = 0
    for item in worked_list:
        if print_stage_on:
            item_idx += 1
            msg = f'Всего: {item_lenght} ккн, - {item_idx}'
            print(msg, end='\r' )
        kkn_chars = item.get("kkn_chars", [])
        chars_len = len(kkn_chars)
        # Записываем основные поля (в 1-й строке записи)
        for key, col in main_keys.items():
            val = item.get(key, "")
            # Если значение не строка, преобразуем в строку
            if not isinstance(val, str):
                val = str(val)
            ws.cell(row=current_row, column=col, value=val)

        is_rus_str = "Да" if item['is_russian'] else ""
        ws.cell(row=current_row, column=18, value=is_rus_str)

        for i in range(chars_len):
            char = kkn_chars[i]
            row_i = current_row + i
            belongs_to_kkn = char.get('belongs_to_kkn')
            if belongs_to_kkn:

                char_name_cell = ws.cell(row=row_i, column=6, value=char.get("name", ""))

                vr_min = ""
                vr_max = ""
                value_str_9 = ""
                value_str_10 = ""
                if char.get("value_is_range"):
                    vr_min, vr_max = char.get("value_range", (None, None))
                    # vr_min = str(vr_min)
                    # vr_max = str(vr_max)

                else:
                    values = char.get("values", [])
                    if len(values) > 1:
                        value_str_9 = "; ".join(values)
                    elif len(values) == 0:
                        print(item)
                    else:
                        value_str_10 = values[0]

                ws.cell(row=row_i, column=7, value=vr_min)
                ws.cell(row=row_i, column=8, value=vr_max)
                ws.cell(row=row_i, column=9, value=value_str_9)
                ws.cell(row=row_i, column=10, value=value_str_10)
                unit_str = char.get("unit", "-")
                ws.cell(row=row_i, column=11, value=unit_str).alignment = Alignment(horizontal='center')

                ws.cell(row=row_i, column=kkn_chars_start_col + 9, value=item.get("product_part", ""))
                date_datetime = item.get("upt_date", '')

                ws.cell(row=row_i, column=kkn_chars_start_col + 10, value=date_datetime).number_format = 'DD.MM.YYYY'

                # ДАННЫЕ ИЗ КТРУ
                ktru_char = char['ktru_data']['ktru_char_origin']
                if not ktru_char:
                    # в ктру нет характеристики из ККН
                    # характеристика не из КТРУ
                    continue
                values = "; ".join(ktru_char['values'])
                ws.cell(row=row_i, column=kkn_chars_start_col + 15, value=values)
                ktru_unit_str = ktru_char['unit'] if ktru_char['unit'] else '-'
                ws.cell(row=row_i, column=kkn_chars_start_col + 16, value=ktru_unit_str)

                if ktru_char['isRequired']:
                    ws.cell(row=row_i, column=kkn_chars_start_col + 13, value=ktru_char['name']).fill = orange_fill
                else:
                    ws.cell(row=row_i, column=kkn_chars_start_col + 13, value=ktru_char['name'])

                # (8 вариантов исходов 4*1*2)
                if ktru_char['version_data']['isNew']: # Используется + новая
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Новая').fill = green_fill
                    # значения совпали - совпадение маловероятно
                    if char["ktru_data"]['all_right']: # Используется + новая + значения совпали
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value='').fill = green_fill
                    else:  # Используется + новая + не совпали значения
                        # столб с пропущ значениями
                        missing_values = '; '.join(char["ktru_data"]['errors']) if char["ktru_data"]['errors'] else ''
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value=missing_values).fill = red_fill

                elif ktru_char['version_data']['isDelete']: # Используется + удалена
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Удаленная').fill = red_fill
                    char_name_cell.fill = yellow_fill
                    if char["ktru_data"]['all_right']: # Используется + удалена + значения совпали
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value='').fill = green_fill
                    else:
                        missing_values = '; '.join(char["ktru_data"]['errors']) if char["ktru_data"]['errors'] else ''
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value=missing_values).fill = red_fill

                elif ktru_char['version_data']['isChanged']: # Используется + изменена
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Измененная').fill = orange_fill
                    # добавленные значения
                    added_ktru_values = ';'.join(ktru_char['version_data']['new_val'])
                    ws.cell(row=row_i, column=kkn_chars_start_col + 18, value=added_ktru_values).fill = green_fill
                    # удаленные значения
                    deleted_ktru_values = ';'.join(ktru_char['version_data']['delete_val'])
                    ws.cell(row=row_i, column=kkn_chars_start_col + 19, value=deleted_ktru_values).fill = red_fill

                    if char["ktru_data"]['all_right']: # Используется + изменена + значения совпали
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value='').fill = green_fill
                    else:
                        missing_values = '; '.join(char["ktru_data"]['errors']) if char["ktru_data"]['errors'] else ''
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value=missing_values).fill = red_fill
                        char_name_cell.fill = yellow_fill

                else: # Используется + старая
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Старая')
                    if char["ktru_data"]['all_right']: # Используется + старая + значения совпали
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value='').fill = green_fill
                    else:
                        missing_values = '; '.join(char["ktru_data"]['errors']) if char["ktru_data"]['errors'] else ''
                        char_name_cell.fill = yellow_fill
                        ws.cell(row=row_i, column=kkn_chars_start_col + 14, value=missing_values).fill = red_fill

            else:
                char_name_cell = char_name_cell = ws.cell(row=row_i, column=6, value='')
                for c_offset in range(6):
                    ws.cell(row=row_i, column=kkn_chars_start_col + c_offset, value="")
                ws.cell(row=row_i, column=kkn_chars_start_col + 9, value=item.get("product_part", ""))

                # ДАННЫЕ ИЗ КТРУ
                ktru_char = char['ktru_data']['ktru_char_origin']

                values = "; ".join(ktru_char['values'])
                ws.cell(row=row_i, column=kkn_chars_start_col + 15, value=values)
                ktru_unit_str = ktru_char['unit'] if ktru_char['unit'] else '-'
                ws.cell(row=row_i, column=kkn_chars_start_col + 16, value=ktru_unit_str)
                if ktru_char['isRequired']:
                    ws.cell(row=row_i, column=kkn_chars_start_col + 13, value=ktru_char['name']).fill = orange_fill
                    char_name_cell.fill = yellow_fill
                else:
                    ws.cell(row=row_i, column=kkn_chars_start_col + 13, value=ktru_char['name'])

                if ktru_char['version_data']['isNew']: # новая
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Новая').fill = green_fill
                    char_name_cell.fill = yellow_fill
                elif ktru_char['version_data']['isDelete']: # удалена
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Удаленная').fill = red_fill
                elif ktru_char['version_data']['isChanged']: # изменена
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Измененная').fill = orange_fill
                    # добавленные значения
                    added_ktru_values = '; '.join(ktru_char['version_data']['new_val'])
                    ws.cell(row=row_i, column=kkn_chars_start_col + 18, value=added_ktru_values).fill = green_fill
                    # удаленные значения
                    deleted_ktru_values = '; '.join(ktru_char['version_data']['delete_val'])
                    ws.cell(row=row_i, column=kkn_chars_start_col + 19, value=deleted_ktru_values).fill = red_fill
                else: # старая
                    ws.cell(row=row_i, column=kkn_chars_start_col + 17, value='Старая')

        # Объединяем ячейки с основными полями по высоте, равной количеству строк характеристик
        for col in merged_cals:
            ws.merge_cells(start_row=current_row, start_column=col,
                           end_row=current_row + chars_len-1, end_column=col)
            ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Переходим к следующей записи
        current_row += chars_len

    # устанавливаем границы и параметры шрифта всей таблицы
    font = Font(name='Times New Roman', size=10)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=25):
        for cell in row:
            cell.border = border
            cell.font = font

    wb.save(updated_ktru_excel_path)
