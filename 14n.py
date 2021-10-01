'''Замена 12n.py
https://tokmakov.msk.ru/blog/item/71 - некоторые моменты уже изменились
'''
import openpyxl
import time

start_time = time.time()

exel_name_ais = 'C:/Users/G.Tishchenko/Desktop/ais.xlsx'
exel_name_kkn = 'C:/Users/G.Tishchenko/Desktop/49_kkn.xlsx'

# файл где список ккн, которые меняли
name_my_file = 'C:/Users/G.Tishchenko/Desktop/sentember_fresh.xlsx'

test_ais = 'C:/Users/G.Tishchenko/Desktop/test_ais.xlsx'
test_kkn = 'C:/Users/G.Tishchenko/Desktop/test_kkn.xlsx'


def import_dicts_from_excel(file_name):
    counter_kkn = 0
    counter_char = 0
    # Открываем файл только на чтение - по быстрому
    wb = openpyxl.load_workbook(file_name, read_only = True, data_only = True)

    # Смотрим сколько в файле листов
    listofsheets = wb.sheetnames

    #active_sheet = wb[listofsheets[0]]

    #all_row = active_sheet['1']

    pre_name = str()
    kkn_dict = {}

    for current_sheet in listofsheets:
        active_sheet = wb[current_sheet]
    # все строки на текущем листе листе
        for row in active_sheet.rows: #

            row_list = []
            # чтение ячеейк в строках и добавление в ""список строки"
            for cell in row[:12]:
                row_list.append(cell.value)

            if type(row_list[1]) == str:
                # когда начинается новый ккн
                pre_name = row_list[1]

                # Если в характеристике есть изменяемые значения
                if type(row_list[8]) == str:
                    temp_list = row_list[8].split(';')
                    # Удаляем пробелы из элементов
                    for i in range(len(temp_list)):
                        temp_list[i] = temp_list[i].lstrip().rstrip()
                    value_set = set(temp_list)
                    char_dict = {row_list[5]:[row_list[6], row_list[7], value_set, row_list[9], row_list[10]]}
                else:
                    char_dict = {row_list[5]:[row_list[6], row_list[7], row_list[8], row_list[9], row_list[10]]}

                # создается словарь для ккн-а
                kkn = {row_list[1]:[row_list[2], row_list[3], row_list[4], char_dict, row_list[11]]} #row_list[0],

                # строка из excel ввиде словаря добавляется в общий словарь (создается новый ключ)
                kkn_dict = kkn_dict | kkn
                counter_kkn += 1
                counter_char += 1
            else:
                # когда продолжается новый ккн

                # Если в характеристике есть изменяемые значения
                if type(row_list[8]) == str:
                    temp_list = row_list[8].split(';')
                    # Удаляем пробелы из элементов
                    for i in range(len(temp_list)):
                        temp_list[i] = temp_list[i].lstrip().rstrip()
                    value_set = set(temp_list)
                    char_dict = {row_list[5]:[row_list[6], row_list[7], value_set, row_list[9], row_list[10]]}
                    #kkn_dict[pre_name[4]] = kkn_dict[pre_name[4]] | char_dict
                else:
                    char_dict = {row_list[5]:[row_list[6], row_list[7], row_list[8], row_list[9], row_list[10]]}
                # строка из excel ввиде словаря добавляется в общий словарь (добавляется к сущ. по ключу)
                kkn_dict[pre_name][4] = kkn_dict[pre_name][4] | char_dict
                counter_char += 1

    print(f'Колво ккн-в:  {counter_kkn - 3}, Колво характ-к (строк):  {counter_char- 2}')
    # тут отняли первые строки
    return kkn_dict



kkn_dict_from_ais = import_dicts_from_excel(exel_name_ais)

kkn_dict_from_kkn = import_dicts_from_excel(exel_name_kkn)


#test_from_ais = import_dicts_from_excel(test_ais)
#test_from_kkn = import_dicts_from_excel(test_kkn)

cur_sec = round((time.time() - start_time), 2)
print(f'Вревмя выполнения: {int(cur_sec // 60)} мин. {cur_sec} сек.)')


def search_set(file_name):
    wb = openpyxl.load_workbook(file_name, read_only = True, data_only = True)
    active_sheet = wb['names']

    search_val = set()

    for row in active_sheet.rows: #
        # чтение ячеейк в строках и добавление в ""список строки"
        for cell in row[:1]:
            #print(cell.value)

            search_val.update({cell.value})
    return search_val


some_set = search_set(name_my_file)

row = 0
for kkn_name in some_set:
    #print(kkn_name)
    row+=1
    if kkn_name in kkn_dict_from_ais:
        print(row)
        for all_char_index in range(len(kkn_dict_from_ais[kkn_name])):
            if type(kkn_dict_from_ais[kkn_name][all_char_index]) != dict:
                if kkn_dict_from_ais[kkn_name][all_char_index] != kkn_dict_from_kkn[kkn_name][all_char_index]:
                    print(kkn_name)
                    print(f'В первом: "{kkn_dict_from_ais[kkn_name][all_char_index]}", во втором: "{kkn_dict_from_kkn[kkn_name][all_char_index]}"')
