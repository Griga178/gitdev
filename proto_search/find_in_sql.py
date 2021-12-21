import sqlite3
import openpyxl
# https://docs.microsoft.com/ru-ru/sql/t-sql/queries/where-transact-sql?view=sql-server-ver15
'''
Поиск контрактов по совпадениям
входящие
название товара
окпд
цена
'''

con = sqlite3.connect('product_base.db')
cur = con.cursor()

def search_in_base(product_name = None, product_okpd = None, mid_price = None):
    all_excecute = f'SELECT * FROM products WHERE product_name LIKE  "%{product_name}%" and product_country LIKE "%РОССИЯ%"' # убрать РОССИЯ ДОБАВИТЬ (
    if mid_price:
        max_price = mid_price * 1.2
        min_price = mid_price * 0.8
        third_part = f'and price Between {min_price} and {max_price}'
    else:
        third_part = ''
    if product_okpd:
        second_part = f'or okpd_num = "{product_okpd}")' # заменене or...) - and
    else:
        second_part = ''
    execute_message = all_excecute + second_part + third_part
    sql_answer = cur.execute(execute_message)
    return sql_answer


def excle_parser(file_name):
    common_list = []
    wb = openpyxl.load_workbook(file_name, read_only = True, data_only = True)
    active_sheet = wb['list']
    for row in active_sheet.rows:
        # Читаем каждую строку
        row_list = []
        for cell in row:
            # добавляем 12 ячеек строки в "список"
            #my_cell = str(cell.value)
            my_cell = cell.value
            row_list.append(my_cell)
        #print(row_list)
        common_list.append([row_list[1], row_list[3].strip(), row_list[0], row_list[2]])

    #for el in common_list:
    #    print(el)
    return common_list

file_name = 'C:/Users/G.Tishchenko/Desktop/kkn_nup.xlsx'
#common_list = excle_parser(file_name)


def find_proto(common_list):
    '''
    Возвращает список [[kkn_name, (sql_answers)], [kkn_name2, (sql_answers)]...]
    '''
    sql_answers_count = 0
    common_list_tosave = []
    for element in common_list:
        #gov_list.append
        #print(element[0])
        rows = search_in_base(element[1], element[2], float(element[3]))

        for el in rows:
            if el:
                #print(element[0])
                sql_answers_count += 1
                #print(el)
                common_list_tosave.append([element[0], el])


    print('найдено ', sql_answers_count)

    return common_list_tosave

#excel_llist = find_proto(common_list)



def print_to_excel():
    gov_adress_page = 'https://zakupki.gov.ru/epz/contract/contractCard/common-info.html?reestrNumber='
    excle_name = 'C:/Users/G.Tishchenko/Desktop/Monoblock.xlsx'
    wb = openpyxl.Workbook()
    sheet_name = wb.create_sheet("checked2")
    active_sheet = wb['checked2']
    row_count =  0
    for el in excel_llist:
        row_count += 1
        #print(el[0])
        print(el[1][1], el[1][8], el[1][2], el[1][11])
        active_sheet.cell(row = row_count, column = 1).value = el[0] # Имя ККН
        active_sheet.cell(row = row_count, column = 2).value = el[1][1] # Имя в закупках
        active_sheet.cell(row = row_count, column = 3).value = el[1][8] # Цена
        active_sheet.cell(row = row_count, column = 4).value = el[1][2] # Страна
        active_sheet.cell(row = row_count, column = 5).value = gov_adress_page + el[1][11] # номер контракта
    wb.save(excle_name)

#print_to_excel()

sql_list = search_in_base('Моноблок')

gov_adress_page = 'https://zakupki.gov.ru/epz/contract/contractCard/common-info.html?reestrNumber='
excle_name = 'C:/Users/G.Tishchenko/Desktop/Monoblock.xlsx'
wb = openpyxl.Workbook()
sheet_name = wb.create_sheet("checked2")
active_sheet = wb['checked2']
row_count =  0

for el in sql_list:
    row_count += 1
    column_num = 1
    #print(el)
    for jey in el:
        active_sheet.cell(row = row_count, column = column_num).value = jey # Имя ККН
        column_num += 1
wb.save(excle_name)
con.close()
