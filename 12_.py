from openpyxl import Workbook, load_workbook
import os
import shutil

def read_file(name: str, sheet_name: str = 'Лист1') -> list[dict]:
    wb = load_workbook(name, read_only = True, data_only = True)
    active_sheet = wb[sheet_name]
    return_list = []
    for row in active_sheet.iter_rows(min_row = 2, values_only = True):

        return_list.append(
            {
            'old_name': row[1],
            'new_name': row[3],
            }
        )

    return return_list

'''
ООО ЧЕЕЕНЬ СЫЫЫРООО

МЕНЯЕМ НАЗВАНИЯ ФАЙЛОВ ПО СПИСКУ ИЗ EXCEL
ОПИСАНИЕ

1
2    'Номер скрина OLD',
3    'Номер скрина NEW',
4

'''
SCR_FOLDER = 'C:/Users/G.Tishchenko/Desktop/screenCap/re/'
XL_FILE = 'C:/Users/G.Tishchenko/Desktop/3 кв 2024/t.xlsx'




names = read_file(XL_FILE, "Лист2")

print(len(names))
# создаем временную папку
temp_folder = "temp"
os.mkdir(SCR_FOLDER + temp_folder)

# Перемещаем в нее файлы с новым названием
for n in names:

    try:
        old_name = SCR_FOLDER + str(n['old_name']) + '.jpg'
        new_name = SCR_FOLDER + temp_folder + '/' + str(n['new_name']) + '.jpg'
        shutil.copy(old_name, new_name)

    except:
        print(f'Er: {n["old_name"]} --> {n["new_name"]}')
        print(old_name, new_name)
